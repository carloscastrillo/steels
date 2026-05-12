from pathlib import Path
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = BASE_DIR / "db" / "steel_mvp.db"
OUTPUT_PATH = BASE_DIR / "exports" / "sourcing_report.xlsx"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF5FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def style_header(ws, row_number: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_label(cell):
    cell.font = Font(bold=True)
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER


def style_data_block(ws, start_row: int, end_row: int, max_col: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def autofit_like(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))

    for col_idx, width in widths.items():
        adjusted = min(max(width + 2, 10), 35)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def build_executive_summary_sheet(conn, wb: Workbook):
    ws = wb.active
    ws.title = "executive_summary"

    summary = conn.execute("""
        SELECT
            COUNT(*) AS total_requests,
            COALESCE(SUM(sr.requested_tons), 0) AS total_tons,
            COALESCE(SUM(CASE WHEN srs.best_option_code IS NOT NULL THEN 1 ELSE 0 END), 0) AS requests_with_best_option,
            COALESCE(SUM(CASE WHEN srs.savings_total_vs_am_spot IS NOT NULL THEN srs.savings_total_vs_am_spot ELSE 0 END), 0) AS total_savings_vs_am_spot,
            AVG(CASE WHEN srs.savings_total_vs_am_spot IS NOT NULL THEN srs.savings_total_vs_am_spot END) AS avg_savings_per_request,
            AVG(CASE WHEN srs.delta_best_vs_am_spot IS NOT NULL THEN srs.delta_best_vs_am_spot END) AS avg_delta_eur_per_ton
        FROM sourcing_request_shortlist srs
        JOIN sourcing_requests sr ON sr.id = srs.sourcing_request_id
    """).fetchone()

    total_quotes = conn.execute("""
        SELECT COUNT(*) AS total_quotes
        FROM sourcing_quotes
    """).fetchone()["total_quotes"]

    total_decisions = conn.execute("""
        SELECT COUNT(*) AS total_decisions
        FROM sourcing_decisions
    """).fetchone()["total_decisions"]

    total_selected_spend = conn.execute("""
        SELECT COALESCE(SUM(q.total_estimated_cost), 0) AS total_selected_spend
        FROM sourcing_decisions d
        JOIN sourcing_quotes q ON q.id = d.selected_quote_id
    """).fetchone()["total_selected_spend"]

    staging = conn.execute("""
        SELECT
            COUNT(*) AS total_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'pending' THEN 1 ELSE 0 END), 0) AS pending_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'approved' THEN 1 ELSE 0 END), 0) AS approved_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'rejected' THEN 1 ELSE 0 END), 0) AS rejected_staging_quotes
        FROM stg_supplier_quotes
    """).fetchone()

    awarded_suppliers = conn.execute("""
        SELECT
            q.supplier_code,
            q.supplier_name,
            COUNT(*) AS wins,
            COALESCE(SUM(q.total_estimated_cost), 0) AS total_awarded
        FROM sourcing_decisions d
        JOIN sourcing_quotes q ON q.id = d.selected_quote_id
        GROUP BY q.supplier_code, q.supplier_name
        ORDER BY wins DESC, total_awarded DESC
        LIMIT 10
    """).fetchall()

    top_requests = conn.execute("""
        SELECT
            srs.sourcing_request_id,
            sr.our_ref,
            c.name AS client_name,
            rs.product,
            rs.grade,
            sr.requested_tons,
            srs.best_option_code,
            srs.best_supplier_name,
            srs.best_total_cost,
            srs.savings_total_vs_am_spot
        FROM sourcing_request_shortlist srs
        JOIN sourcing_requests sr ON sr.id = srs.sourcing_request_id
        JOIN clients c ON c.id = sr.client_id
        JOIN request_specs rs ON rs.id = sr.request_spec_id
        ORDER BY srs.savings_total_vs_am_spot DESC
        LIMIT 10
    """).fetchall()

    ws["A1"] = "EXECUTIVE SUMMARY"
    style_section_label(ws["A1"])

    metrics = [
        ("Total requests", summary["total_requests"]),
        ("Total tons", summary["total_tons"]),
        ("Requests con mejor opción", summary["requests_with_best_option"]),
        ("Total sourcing quotes", total_quotes),
        ("Total decisions", total_decisions),
        ("Total selected spend", total_selected_spend),
        ("Ahorro total potencial vs AM_SPOT", summary["total_savings_vs_am_spot"]),
        ("Ahorro medio por request", summary["avg_savings_per_request"]),
        ("Delta medio EUR/t vs AM_SPOT", summary["avg_delta_eur_per_ton"]),
        ("Total staging quotes", staging["total_staging_quotes"]),
        ("Pending staging quotes", staging["pending_staging_quotes"]),
        ("Approved staging quotes", staging["approved_staging_quotes"]),
        ("Rejected staging quotes", staging["rejected_staging_quotes"]),
    ]

    row = 3
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    for r in range(3, row):
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="TOP AWARDED SUPPLIERS")
    style_section_label(ws.cell(row=row, column=1))
    row += 1

    supplier_headers = ["supplier_code", "supplier_name", "wins", "total_awarded"]
    for col_idx, header in enumerate(supplier_headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, len(supplier_headers))

    start_supplier_row = row + 1
    for supplier in awarded_suppliers:
        row += 1
        ws.cell(row=row, column=1, value=supplier["supplier_code"])
        ws.cell(row=row, column=2, value=supplier["supplier_name"])
        ws.cell(row=row, column=3, value=supplier["wins"])
        ws.cell(row=row, column=4, value=supplier["total_awarded"])

    if row >= start_supplier_row:
        style_data_block(ws, start_supplier_row, row, len(supplier_headers))
        for r in range(start_supplier_row, row + 1):
            ws.cell(row=r, column=4).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="TOP REQUESTS BY SAVINGS")
    style_section_label(ws.cell(row=row, column=1))
    row += 1

    request_headers = [
        "sourcing_request_id",
        "our_ref",
        "client_name",
        "product",
        "grade",
        "requested_tons",
        "best_option_code",
        "best_supplier_name",
        "best_total_cost",
        "savings_total_vs_am_spot",
    ]
    for col_idx, header in enumerate(request_headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, len(request_headers))

    start_request_row = row + 1
    for req in top_requests:
        row += 1
        ws.cell(row=row, column=1, value=req["sourcing_request_id"])
        ws.cell(row=row, column=2, value=req["our_ref"])
        ws.cell(row=row, column=3, value=req["client_name"])
        ws.cell(row=row, column=4, value=req["product"])
        ws.cell(row=row, column=5, value=req["grade"])
        ws.cell(row=row, column=6, value=req["requested_tons"])
        ws.cell(row=row, column=7, value=req["best_option_code"])
        ws.cell(row=row, column=8, value=req["best_supplier_name"])
        ws.cell(row=row, column=9, value=req["best_total_cost"])
        ws.cell(row=row, column=10, value=req["savings_total_vs_am_spot"])

    if row >= start_request_row:
        style_data_block(ws, start_request_row, row, len(request_headers))
        for r in range(start_request_row, row + 1):
            ws.cell(row=r, column=6).number_format = '#,##0.00'
            ws.cell(row=r, column=9).number_format = '#,##0.00'
            ws.cell(row=r, column=10).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)


def build_shortlist_sheet(conn, wb: Workbook):
    ws = wb.create_sheet("shortlist_requests")

    headers = [
        "sourcing_request_id",
        "our_ref",
        "client_name",
        "product",
        "grade",
        "thickness_mm",
        "width_mm",
        "cw_min",
        "cw_max",
        "requested_tons",
        "best_option_code",
        "best_supplier_name",
        "best_unit_cost",
        "best_total_cost",
        "second_option_code",
        "second_unit_cost",
        "third_option_code",
        "third_unit_cost",
        "am_spot_unit_cost",
        "delta_best_vs_am_spot",
        "savings_total_vs_am_spot",
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    rows = conn.execute("""
        SELECT
            srs.sourcing_request_id,
            sr.our_ref,
            c.name AS client_name,
            rs.product,
            rs.grade,
            rs.thickness_mm,
            rs.width_mm,
            rs.cw_min,
            rs.cw_max,
            sr.requested_tons,
            srs.best_option_code,
            srs.best_supplier_name,
            srs.best_unit_cost,
            srs.best_total_cost,
            srs.second_option_code,
            srs.second_unit_cost,
            srs.third_option_code,
            srs.third_unit_cost,
            srs.am_spot_unit_cost,
            srs.delta_best_vs_am_spot,
            srs.savings_total_vs_am_spot
        FROM sourcing_request_shortlist srs
        JOIN sourcing_requests sr ON sr.id = srs.sourcing_request_id
        JOIN clients c ON c.id = sr.client_id
        JOIN request_specs rs ON rs.id = sr.request_spec_id
        ORDER BY srs.sourcing_request_id
    """).fetchall()

    start_row = 2
    for row in rows:
        ws.append(list(row))
    end_row = ws.max_row

    if end_row >= start_row:
        style_data_block(ws, start_row, end_row, len(headers))

    currency_cols = [13, 14, 16, 18, 19, 20, 21]
    for col in currency_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)


def build_summary_sheet(conn, wb: Workbook):
    ws = wb.create_sheet("summary")

    summary = conn.execute("""
        SELECT
            COUNT(*) AS total_requests,
            SUM(sr.requested_tons) AS total_tons,
            SUM(CASE WHEN srs.best_option_code IS NOT NULL THEN 1 ELSE 0 END) AS requests_with_best_option,
            SUM(CASE WHEN srs.savings_total_vs_am_spot IS NOT NULL THEN srs.savings_total_vs_am_spot ELSE 0 END) AS total_savings_vs_am_spot,
            AVG(CASE WHEN srs.savings_total_vs_am_spot IS NOT NULL THEN srs.savings_total_vs_am_spot END) AS avg_savings_per_request,
            AVG(CASE WHEN srs.delta_best_vs_am_spot IS NOT NULL THEN srs.delta_best_vs_am_spot END) AS avg_delta_eur_per_ton
        FROM sourcing_request_shortlist srs
        JOIN sourcing_requests sr ON sr.id = srs.sourcing_request_id
    """).fetchone()

    winners = conn.execute("""
        SELECT
            best_option_code,
            best_supplier_name,
            COUNT(*) AS wins,
            SUM(savings_total_vs_am_spot) AS total_savings,
            AVG(savings_total_vs_am_spot) AS avg_savings,
            AVG(delta_best_vs_am_spot) AS avg_delta_eur_per_ton
        FROM sourcing_request_shortlist
        WHERE best_option_code IS NOT NULL
        GROUP BY best_option_code, best_supplier_name
        ORDER BY wins DESC, total_savings DESC
    """).fetchall()

    ws["A1"] = "RESUMEN GLOBAL"
    style_section_label(ws["A1"])

    metrics = [
        ("Total requests", summary["total_requests"]),
        ("Total tons", summary["total_tons"]),
        ("Requests con mejor opción", summary["requests_with_best_option"]),
        ("Ahorro total potencial vs AM_SPOT", summary["total_savings_vs_am_spot"]),
        ("Ahorro medio por request", summary["avg_savings_per_request"]),
        ("Delta medio EUR/t vs AM_SPOT", summary["avg_delta_eur_per_ton"]),
    ]

    row = 3
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    for r in range(3, 9):
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="GANADORES POR PROVEEDOR / OPCIÓN")
    style_section_label(ws.cell(row=row, column=1))
    row += 1

    winner_headers = [
        "best_option_code",
        "best_supplier_name",
        "wins",
        "total_savings",
        "avg_savings",
        "avg_delta_eur_per_ton",
    ]
    for col_idx, header in enumerate(winner_headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, len(winner_headers))

    start_winner_row = row + 1
    for winner in winners:
        row += 1
        for col_idx, value in enumerate(winner, start=1):
            ws.cell(row=row, column=col_idx, value=value)

    if row >= start_winner_row:
        style_data_block(ws, start_winner_row, row, len(winner_headers))
        for r in range(start_winner_row, row + 1):
            ws.cell(row=r, column=4).number_format = '#,##0.00'
            ws.cell(row=r, column=5).number_format = '#,##0.00'
            ws.cell(row=r, column=6).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)


def main():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"No existe la base de datos: {DB_PATH}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row

        wb = Workbook()
        build_executive_summary_sheet(conn, wb)
        build_shortlist_sheet(conn, wb)
        build_summary_sheet(conn, wb)
        wb.save(OUTPUT_PATH)

    print("Exportación completada correctamente.")
    print(f"Archivo generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()