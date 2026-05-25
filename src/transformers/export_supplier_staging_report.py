from __future__ import annotations

from datetime import datetime
from pathlib import Path
import os
import re
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = Path(os.environ.get("STEEL_DB_PATH", BASE_DIR / "db" / "steel_mvp.db"))
EXPORTS_DIR = BASE_DIR / "exports"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF5FB")
APPROVAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


HEADERS = [
    "id",
    "supplier_code",
    "supplier_name",
    "documento",
    "coating_raw",
    "extracted_grade",
    "extracted_thickness_mm",
    "extracted_width_mm",
    "extracted_price_per_ton",
    "currency",
    "review_status",
    "needs_manual_review",
    "notes",
    "created_at",
    "raw_text_snippet",
    "APROBACION_OPERADOR",
]


def table_has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(row[1] == column for row in rows)


def safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name.strip()) or "UNKNOWN"
    cleaned = cleaned[:31]

    base = cleaned
    i = 2
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = f"{base[:31-len(suffix)]}{suffix}"
        i += 1

    used.add(cleaned)
    return cleaned


def fetch_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    coating_expr = "q.coating_raw" if table_has_column(conn, "stg_supplier_quotes", "coating_raw") else "NULL"

    query = f"""
        SELECT
            q.id,
            q.supplier_code,
            q.supplier_name,
            COALESCE(d.file_name, '') AS documento,
            {coating_expr} AS coating_raw,
            q.extracted_grade,
            q.extracted_thickness_mm,
            q.extracted_width_mm,
            q.extracted_price_per_ton,
            q.currency,
            q.review_status,
            q.needs_manual_review,
            q.notes,
            q.created_at,
            q.raw_text_snippet
        FROM stg_supplier_quotes q
        LEFT JOIN stg_supplier_documents d
            ON d.id = q.supplier_document_id
        WHERE COALESCE(q.review_status, 'pending') = 'pending'
        ORDER BY q.supplier_code, d.file_name, q.id
    """

    return conn.execute(query).fetchall()


def group_by_supplier(rows: list[sqlite3.Row]) -> dict[str, list[sqlite3.Row]]:
    grouped: dict[str, list[sqlite3.Row]] = {}

    for row in rows:
        supplier_code = row["supplier_code"] or "UNKNOWN"
        grouped.setdefault(supplier_code, []).append(row)

    return grouped


def style_header(ws, row_number: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    approval_cell = ws.cell(row=row_number, column=max_col)
    approval_cell.fill = APPROVAL_FILL


def style_data(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.cell(row=row, column=max_col).fill = APPROVAL_FILL


def autofit_like(ws) -> None:
    max_width_by_col = {
        "A": 10,
        "B": 16,
        "C": 24,
        "D": 38,
        "E": 24,
        "F": 42,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 12,
        "K": 16,
        "L": 20,
        "M": 50,
        "N": 22,
        "O": 50,
        "P": 24,
    }

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))

        width = min(max(max_len + 2, 10), max_width_by_col.get(letter, 35))
        ws.column_dimensions[letter].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24


def add_validation(ws, start_row: int, end_row: int, approval_col: int) -> None:
    if end_row < start_row:
        return

    col_letter = get_column_letter(approval_col)
    dv = DataValidation(
        type="list",
        formula1='"approved,rejected"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Valor permitido: approved o rejected"
    dv.errorTitle = "Valor no válido"
    dv.prompt = "Selecciona approved o rejected"
    dv.promptTitle = "Aprobación operador"

    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def build_summary_sheet(wb: Workbook, grouped: dict[str, list[sqlite3.Row]]) -> None:
    ws = wb.active
    ws.title = "SUMMARY"

    ws["A1"] = "Supplier staging review report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = SECTION_FILL

    ws.append([])
    ws.append(["supplier_code", "pending_quotes"])

    style_header(ws, 3, 2)

    for supplier_code, rows in grouped.items():
        ws.append([supplier_code, len(rows)])

    if ws.max_row >= 4:
        style_data(ws, 4, ws.max_row, 2)

    ws.freeze_panes = "A4"
    autofit_like(ws)


def build_supplier_sheet(wb: Workbook, supplier_code: str, rows: list[sqlite3.Row], used_names: set[str]) -> None:
    sheet_name = safe_sheet_name(supplier_code, used_names)
    ws = wb.create_sheet(sheet_name)

    ws.append(HEADERS)
    style_header(ws, 1, len(HEADERS))

    for row in rows:
        ws.append([
            row["id"],
            row["supplier_code"],
            row["supplier_name"],
            row["documento"],
            row["coating_raw"],
            row["extracted_grade"],
            row["extracted_thickness_mm"],
            row["extracted_width_mm"],
            row["extracted_price_per_ton"],
            row["currency"],
            row["review_status"],
            row["needs_manual_review"],
            row["notes"],
            row["created_at"],
            row["raw_text_snippet"],
            "",
        ])

    if ws.max_row >= 2:
        style_data(ws, 2, ws.max_row, len(HEADERS))
        add_validation(ws, 2, ws.max_row, len(HEADERS))

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=7).number_format = "0.000"
        ws.cell(row=row_idx, column=8).number_format = "0.000"
        ws.cell(row=row_idx, column=9).number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_like(ws)


def main() -> None:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"No existe la DB: {DB_PATH}")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = fetch_rows(conn)

    grouped = group_by_supplier(rows)

    wb = Workbook()
    build_summary_sheet(wb, grouped)

    used_names = {"SUMMARY"}
    for supplier_code, supplier_rows in grouped.items():
        build_supplier_sheet(wb, supplier_code, supplier_rows, used_names)

    stamp = datetime.now().strftime("%Y-%m-%d")
    output_path = EXPORTS_DIR / f"supplier_staging_report_{stamp}.xlsx"
    wb.save(output_path)

    print("Exportación staging proveedor completada.")
    print(f"Archivo generado: {output_path}")
    print(f"Quotes pendientes exportadas: {len(rows)}")
    print("Proveedores incluidos:")
    for supplier_code, supplier_rows in grouped.items():
        print(f" - {supplier_code}: {len(supplier_rows)}")


if __name__ == "__main__":
    main()
