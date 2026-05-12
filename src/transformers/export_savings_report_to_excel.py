from pathlib import Path
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = BASE_DIR / "db" / "steel_mvp.db"
OUTPUT_PATH = BASE_DIR / "exports" / "savings_report.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF5FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def style_header(ws, row_number: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_label(cell):
    cell.font = Font(bold=True)
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER


def style_data_block(ws, start_row: int, end_row: int, max_col: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def autofit_like(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(value))

    for col_idx, width in widths.items():
        adjusted = min(max(width + 2, 10), 35)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def fetch_rows(conn):
    decisions = conn.execute("""
        SELECT
            d.id AS decision_id,
            d.sourcing_request_id,
            d.selected_quote_id,
            d.decision_reason,
            d.decided_by,
            d.decided_at,
            sr.our_ref,
            sr.requested_tons,
            sr.status AS request_status,
            c.name AS client_name,
            rs.product,
            rs.grade,
            rs.thickness_mm,
            rs.width_mm,
            q.supplier_code AS selected_supplier_code,
            q.supplier_name AS selected_supplier_name,
            q.total_price_per_ton AS selected_total_price_per_ton,
            q.total_estimated_cost AS selected_total_estimated_cost,
            q.needs_manual_review AS selected_needs_manual_review
        FROM sourcing_decisions d
        JOIN sourcing_requests sr ON sr.id = d.sourcing_request_id
        JOIN clients c            ON c.id = sr.client_id
        JOIN request_specs rs     ON rs.id = sr.request_spec_id
        JOIN sourcing_quotes q    ON q.id = d.selected_quote_id
        ORDER BY d.decided_at DESC, d.id DESC
    """).fetchall()

    result_rows = []

    for row in decisions:
        selected_price = float(row["selected_total_price_per_ton"] or 0)
        requested_tons = float(row["requested_tons"] or 0)

        next_best = conn.execute("""
            SELECT
                supplier_code,
                supplier_name,
                total_price_per_ton
            FROM sourcing_quotes
            WHERE sourcing_request_id = ?
              AND id <> ?
              AND COALESCE(needs_manual_review, 0) = 0
            ORDER BY total_price_per_ton ASC, id ASC
            LIMIT 1
        """, (row["sourcing_request_id"], row["selected_quote_id"])).fetchone()

        excluded_manual_review_count = conn.execute("""
            SELECT COUNT(*)
            FROM sourcing_quotes
            WHERE sourcing_request_id = ?
              AND id <> ?
              AND COALESCE(needs_manual_review, 0) = 1
        """, (row["sourcing_request_id"], row["selected_quote_id"])).fetchone()[0]

        if next_best is not None and next_best["total_price_per_ton"] is not None:
            next_best_supplier = f"{next_best['supplier_code']} | {next_best['supplier_name']}"
            next_best_price = float(next_best["total_price_per_ton"])
            savings_vs_next_best_real_total = (next_best_price - selected_price) * requested_tons
        else:
            next_best_supplier = None
            next_best_price = None
            savings_vs_next_best_real_total = 0.0

        am_spot = conn.execute("""
            SELECT unit_cost
            FROM supplier_options
            WHERE sourcing_request_id = ?
              AND option_code = 'AM_SPOT'
              AND is_comparable = 1
              AND is_rankable = 1
              AND capability_allowed = 1
            LIMIT 1
        """, (row["sourcing_request_id"],)).fetchone()

        if am_spot is not None and am_spot["unit_cost"] is not None:
            am_spot_unit_cost = float(am_spot["unit_cost"])
            savings_vs_am_spot_total = (am_spot_unit_cost - selected_price) * requested_tons
        else:
            am_spot_unit_cost = None
            savings_vs_am_spot_total = 0.0

        result_rows.append({
            "decision_id": row["decision_id"],
            "sourcing_request_id": row["sourcing_request_id"],
            "our_ref": row["our_ref"],
            "client_name": row["client_name"],
            "product": row["product"],
            "grade": row["grade"],
            "thickness_mm": row["thickness_mm"],
            "width_mm": row["width_mm"],
            "requested_tons": requested_tons,
            "selected_supplier_code": row["selected_supplier_code"],
            "selected_supplier_name": row["selected_supplier_name"],
            "selected_total_price_per_ton": selected_price,
            "selected_total_estimated_cost": float(row["selected_total_estimated_cost"] or 0),
            "selected_needs_manual_review": row["selected_needs_manual_review"],
            "next_best_real_supplier": next_best_supplier,
            "next_best_real_price_per_ton": next_best_price,
            "excluded_manual_review_quotes": excluded_manual_review_count,
            "savings_vs_next_best_real_total": savings_vs_next_best_real_total,
            "am_spot_benchmark_per_ton": am_spot_unit_cost,
            "savings_vs_am_spot_total": savings_vs_am_spot_total,
            "decision_reason": row["decision_reason"],
            "decided_by": row["decided_by"],
            "decided_at": row["decided_at"],
            "request_status": row["request_status"],
        })

    return result_rows


def main():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"No existe la base de datos: {DB_PATH}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = fetch_rows(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "savings_report"

    headers = [
        "decision_id",
        "sourcing_request_id",
        "our_ref",
        "client_name",
        "product",
        "grade",
        "thickness_mm",
        "width_mm",
        "requested_tons",
        "selected_supplier_code",
        "selected_supplier_name",
        "selected_total_price_per_ton",
        "selected_total_estimated_cost",
        "selected_needs_manual_review",
        "next_best_real_supplier",
        "next_best_real_price_per_ton",
        "excluded_manual_review_quotes",
        "savings_vs_next_best_real_total",
        "am_spot_benchmark_per_ton",
        "savings_vs_am_spot_total",
        "decision_reason",
        "decided_by",
        "decided_at",
        "request_status",
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    start_row = 2
    for row in rows:
        ws.append([row[h] for h in headers])
    end_row = ws.max_row

    if end_row >= start_row:
        style_data_block(ws, start_row, end_row, len(headers))

    currency_cols = [9, 12, 13, 16, 18, 19, 20]
    for col in currency_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)
    wb.save(OUTPUT_PATH)

    print("Exportación completada correctamente.")
    print(f"Archivo generado: {OUTPUT_PATH}")
    print(f"Filas exportadas: {len(rows)}")


if __name__ == "__main__":
    main()