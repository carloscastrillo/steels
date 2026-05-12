from __future__ import annotations

import argparse
import sqlite3
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent.parent
DB_PATH = BASE_DIR / "db" / "steel_mvp.db"
EXPORTS_DIR = BASE_DIR / "exports"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF5FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera el informe mensual del Steel MVP.")
    parser.add_argument("--month", required=True, help="Mes en formato YYYY-MM, ej: 2026-05")
    return parser.parse_args()


def validate_month(month: str) -> None:
    if len(month) != 7 or month[4] != "-":
        raise ValueError("El parámetro --month debe tener formato YYYY-MM")
    yyyy, mm = month.split("-")
    if not (yyyy.isdigit() and mm.isdigit()):
        raise ValueError("El parámetro --month debe tener formato YYYY-MM")
    m = int(mm)
    if m < 1 or m > 12:
        raise ValueError("Mes no válido en --month")


def style_header(ws, row_number: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_number, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_label(cell):
    cell.font = Font(bold=True)
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER


def style_data_block(ws, start_row: int, end_row: int, max_col: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def autofit_like(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 35)


def run_export_scripts() -> None:
    scripts = [
        BASE_DIR / "src" / "transformers" / "export_sourcing_report_to_excel.py",
        BASE_DIR / "src" / "transformers" / "export_savings_report_to_excel.py",
    ]

    for script in scripts:
        result = subprocess.run([sys.executable, str(script)], cwd=str(BASE_DIR))
        if result.returncode != 0:
            raise RuntimeError(f"Falló la ejecución de {script.name} con código {result.returncode}")


def compute_monthly_metrics(conn: sqlite3.Connection, month: str) -> dict:
    requests_summary = conn.execute("""
        SELECT
            COUNT(*) AS total_requests,
            COALESCE(SUM(requested_tons), 0) AS total_tons
        FROM sourcing_requests
        WHERE substr(sheet_date, 1, 7) = ?
    """, (month,)).fetchone()

    decisions_rows = conn.execute("""
        SELECT
            d.id AS decision_id,
            d.sourcing_request_id,
            d.selected_quote_id,
            d.decided_at,
            sr.our_ref,
            sr.requested_tons,
            c.name AS client_name,
            rs.product,
            rs.grade,
            q.supplier_code,
            q.supplier_name,
            q.total_price_per_ton,
            q.total_estimated_cost
        FROM sourcing_decisions d
        JOIN sourcing_requests sr ON sr.id = d.sourcing_request_id
        JOIN clients c            ON c.id = sr.client_id
        JOIN request_specs rs     ON rs.id = sr.request_spec_id
        JOIN sourcing_quotes q    ON q.id = d.selected_quote_id
        WHERE substr(d.decided_at, 1, 7) = ?
        ORDER BY d.decided_at
    """, (month,)).fetchall()

    decisions_count = len(decisions_rows)
    total_selected_spend = sum(float(r["total_estimated_cost"] or 0) for r in decisions_rows)

    total_savings_vs_am_spot = 0.0
    total_savings_vs_next_best = 0.0
    requests_with_alt_quote = 0

    detailed_request_rows = []

    for r in decisions_rows:
        request_id = r["sourcing_request_id"]
        selected_quote_id = r["selected_quote_id"]
        requested_tons = float(r["requested_tons"] or 0)
        selected_price = float(r["total_price_per_ton"] or 0)

        next_best = conn.execute("""
            SELECT
                supplier_code,
                supplier_name,
                total_price_per_ton
            FROM sourcing_quotes
            WHERE sourcing_request_id = ?
              AND id <> ?
              AND COALESCE(needs_manual_review, 0) = 0
            ORDER BY total_price_per_ton ASC, id ASC
            LIMIT 1
        """, (request_id, selected_quote_id)).fetchone()

        if next_best is not None and next_best["total_price_per_ton"] is not None:
            requests_with_alt_quote += 1
            savings_vs_next_best = (float(next_best["total_price_per_ton"]) - selected_price) * requested_tons
            total_savings_vs_next_best += savings_vs_next_best
            next_best_supplier = f"{next_best['supplier_code']} | {next_best['supplier_name']}"
            next_best_price = float(next_best["total_price_per_ton"])
        else:
            savings_vs_next_best = 0.0
            next_best_supplier = None
            next_best_price = None

        am_spot = conn.execute("""
            SELECT unit_cost
            FROM supplier_options
            WHERE sourcing_request_id = ?
              AND option_code = 'AM_SPOT'
              AND is_comparable = 1
              AND is_rankable = 1
              AND capability_allowed = 1
            LIMIT 1
        """, (request_id,)).fetchone()

        if am_spot is not None and am_spot["unit_cost"] is not None:
            am_spot_unit = float(am_spot["unit_cost"])
            savings_vs_am_spot = (am_spot_unit - selected_price) * requested_tons
            total_savings_vs_am_spot += savings_vs_am_spot
        else:
            am_spot_unit = None
            savings_vs_am_spot = 0.0

        detailed_request_rows.append({
            "decision_id": r["decision_id"],
            "sourcing_request_id": request_id,
            "our_ref": r["our_ref"],
            "client_name": r["client_name"],
            "product": r["product"],
            "grade": r["grade"],
            "supplier_code": r["supplier_code"],
            "supplier_name": r["supplier_name"],
            "selected_total_price_per_ton": selected_price,
            "selected_total_estimated_cost": float(r["total_estimated_cost"] or 0),
            "next_best_supplier": next_best_supplier,
            "next_best_price_per_ton": next_best_price,
            "savings_vs_next_best_total": savings_vs_next_best,
            "am_spot_unit_cost": am_spot_unit,
            "savings_vs_am_spot_total": savings_vs_am_spot,
            "decided_at": r["decided_at"],
        })

    supplier_rows = conn.execute("""
        SELECT
            q.supplier_code,
            q.supplier_name,
            COUNT(*) AS wins,
            COALESCE(SUM(q.total_estimated_cost), 0) AS total_awarded
        FROM sourcing_decisions d
        JOIN sourcing_quotes q ON q.id = d.selected_quote_id
        WHERE substr(d.decided_at, 1, 7) = ?
        GROUP BY q.supplier_code, q.supplier_name
        ORDER BY wins DESC, total_awarded DESC
    """, (month,)).fetchall()

    staging_rows = conn.execute("""
        SELECT
            COUNT(*) AS total_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'pending' THEN 1 ELSE 0 END), 0) AS pending_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'approved' THEN 1 ELSE 0 END), 0) AS approved_staging_quotes,
            COALESCE(SUM(CASE WHEN review_status = 'rejected' THEN 1 ELSE 0 END), 0) AS rejected_staging_quotes
        FROM stg_supplier_quotes
        WHERE substr(created_at, 1, 7) = ?
    """, (month,)).fetchone()

    docs_rows = conn.execute("""
        SELECT
            COUNT(*) AS total_docs,
            COALESCE(SUM(CASE WHEN file_type = 'pdf' THEN 1 ELSE 0 END), 0) AS pdf_docs,
            COALESCE(SUM(n_quotes_extracted), 0) AS quotes_extracted
        FROM stg_supplier_documents
        WHERE substr(imported_at, 1, 7) = ?
    """, (month,)).fetchone()

    return {
        "month": month,
        "requests_count": requests_summary["total_requests"],
        "requests_tons": requests_summary["total_tons"],
        "decisions_count": decisions_count,
        "total_selected_spend": total_selected_spend,
        "total_savings_vs_am_spot": total_savings_vs_am_spot,
        "total_savings_vs_next_best": total_savings_vs_next_best,
        "requests_with_alt_quote": requests_with_alt_quote,
        "supplier_rows": supplier_rows,
        "detailed_request_rows": sorted(
            detailed_request_rows,
            key=lambda x: x["savings_vs_am_spot_total"],
            reverse=True
        ),
        "staging_rows": staging_rows,
        "docs_rows": docs_rows,
    }


def build_overview_sheet(wb: Workbook, metrics: dict):
    ws = wb.active
    ws.title = "overview"

    ws["A1"] = f"MONTHLY REPORT {metrics['month']}"
    style_section_label(ws["A1"])

    rows = [
        ("Requests del mes", metrics["requests_count"]),
        ("Toneladas del mes", metrics["requests_tons"]),
        ("Decisiones del mes", metrics["decisions_count"]),
        ("Spend adjudicado del mes", metrics["total_selected_spend"]),
        ("Ahorro total vs AM_SPOT", metrics["total_savings_vs_am_spot"]),
        ("Ahorro total vs next best", metrics["total_savings_vs_next_best"]),
        ("Requests con alternativa real", metrics["requests_with_alt_quote"]),
        ("Documentos proveedor cargados", metrics["docs_rows"]["total_docs"]),
        ("PDFs proveedor cargados", metrics["docs_rows"]["pdf_docs"]),
        ("Quotes extraídas de documentos", metrics["docs_rows"]["quotes_extracted"]),
        ("Staging quotes del mes", metrics["staging_rows"]["total_staging_quotes"]),
        ("Staging pending", metrics["staging_rows"]["pending_staging_quotes"]),
        ("Staging approved", metrics["staging_rows"]["approved_staging_quotes"]),
        ("Staging rejected", metrics["staging_rows"]["rejected_staging_quotes"]),
    ]

    row = 3
    for label, value in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    for r in range(3, row):
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    row += 2
    ws.cell(row=row, column=1, value="TOP SUPPLIERS DEL MES")
    style_section_label(ws.cell(row=row, column=1))
    row += 1

    headers = ["supplier_code", "supplier_name", "wins", "total_awarded"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
    style_header(ws, row, len(headers))

    start_row = row + 1
    for supplier in metrics["supplier_rows"]:
        row += 1
        ws.cell(row=row, column=1, value=supplier["supplier_code"])
        ws.cell(row=row, column=2, value=supplier["supplier_name"])
        ws.cell(row=row, column=3, value=supplier["wins"])
        ws.cell(row=row, column=4, value=supplier["total_awarded"])

    if row >= start_row:
        style_data_block(ws, start_row, row, len(headers))
        for r in range(start_row, row + 1):
            ws.cell(row=r, column=4).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)


def build_requests_sheet(wb: Workbook, metrics: dict):
    ws = wb.create_sheet("requests")

    headers = [
        "decision_id",
        "sourcing_request_id",
        "our_ref",
        "client_name",
        "product",
        "grade",
        "supplier_code",
        "supplier_name",
        "selected_total_price_per_ton",
        "selected_total_estimated_cost",
        "next_best_supplier",
        "next_best_price_per_ton",
        "savings_vs_next_best_total",
        "am_spot_unit_cost",
        "savings_vs_am_spot_total",
        "decided_at",
    ]

    ws.append(headers)
    style_header(ws, 1, len(headers))

    start_row = 2
    for row in metrics["detailed_request_rows"]:
        ws.append([
            row["decision_id"],
            row["sourcing_request_id"],
            row["our_ref"],
            row["client_name"],
            row["product"],
            row["grade"],
            row["supplier_code"],
            row["supplier_name"],
            row["selected_total_price_per_ton"],
            row["selected_total_estimated_cost"],
            row["next_best_supplier"],
            row["next_best_price_per_ton"],
            row["savings_vs_next_best_total"],
            row["am_spot_unit_cost"],
            row["savings_vs_am_spot_total"],
            row["decided_at"],
        ])

    end_row = ws.max_row
    if end_row >= start_row:
        style_data_block(ws, start_row, end_row, len(headers))
        currency_cols = [9, 10, 12, 13, 14, 15]
        for col in currency_cols:
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=col).number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    autofit_like(ws)


def build_notes_sheet(wb: Workbook, month: str):
    ws = wb.create_sheet("notes")
    ws["A1"] = "NOTES"
    style_section_label(ws["A1"])

    notes = [
        f"Informe mensual generado para: {month}",
        "Este script regenera sourcing_report.xlsx y savings_report.xlsx antes de crear este informe.",
        "Las métricas mensuales de requests se basan en sourcing_requests.sheet_date.",
        "Las métricas mensuales de decisiones se basan en sourcing_decisions.decided_at.",
        "El ahorro vs AM_SPOT se calcula contra supplier_options comparables/rankeables/capability_allowed.",
        "Las quotes staging se contabilizan por stg_supplier_quotes.created_at.",
    ]

    row = 3
    for note in notes:
        ws.cell(row=row, column=1, value=note)
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    autofit_like(ws)


def main():
    args = parse_args()
    validate_month(args.month)

    if not DB_PATH.exists():
        raise FileNotFoundError(f"No existe la base de datos: {DB_PATH}")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    run_export_scripts()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        metrics = compute_monthly_metrics(conn, args.month)

    wb = Workbook()
    build_overview_sheet(wb, metrics)
    build_requests_sheet(wb, metrics)
    build_notes_sheet(wb, args.month)

    output_path = EXPORTS_DIR / f"monthly_report_{args.month}.xlsx"
    wb.save(output_path)

    print("Informe mensual generado correctamente.")
    print(f"Mes: {args.month}")
    print(f"Archivo mensual: {output_path}")
    print(f"Archivo sourcing: {EXPORTS_DIR / 'sourcing_report.xlsx'}")
    print(f"Archivo savings: {EXPORTS_DIR / 'savings_report.xlsx'}")


if __name__ == "__main__":
    main()